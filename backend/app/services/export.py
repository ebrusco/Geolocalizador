import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GOOGLE_TYPE_ES: dict[str, str] = {
    "accounting": "Contabilidad",
    "airport": "Aeropuerto",
    "amusement_park": "Parque de diversiones",
    "aquarium": "Acuario",
    "art_gallery": "Galería de arte",
    "atm": "Cajero automático",
    "bakery": "Panadería",
    "bank": "Banco",
    "bar": "Bar",
    "beauty_salon": "Salón de belleza",
    "bicycle_store": "Bicicletería",
    "book_store": "Librería",
    "bowling_alley": "Bowling",
    "bus_station": "Terminal de buses",
    "cafe": "Café",
    "campground": "Camping",
    "car_dealer": "Concesionaria",
    "car_rental": "Alquiler de autos",
    "car_repair": "Taller mecánico",
    "car_wash": "Lavadero de autos",
    "casino": "Casino",
    "cemetery": "Cementerio",
    "church": "Iglesia",
    "city_hall": "Municipalidad",
    "clothing_store": "Tienda de ropa",
    "convenience_store": "Almacén",
    "courthouse": "Juzgado",
    "dentist": "Dentista",
    "department_store": "Tienda departamental",
    "doctor": "Médico",
    "drugstore": "Farmacia",
    "electrician": "Electricista",
    "electronics_store": "Electrónica",
    "embassy": "Embajada",
    "establishment": "Establecimiento",
    "fire_station": "Bomberos",
    "florist": "Florería",
    "food": "Comida",
    "funeral_home": "Funeraria",
    "furniture_store": "Mueblería",
    "gas_station": "Estación de servicio",
    "gym": "Gimnasio",
    "hair_care": "Peluquería",
    "hardware_store": "Ferretería",
    "health": "Salud",
    "hindu_temple": "Templo hindú",
    "home_goods_store": "Bazar / Hogar",
    "hospital": "Hospital",
    "insurance_agency": "Seguros",
    "jewelry_store": "Joyería",
    "laundry": "Lavandería",
    "lawyer": "Abogado",
    "library": "Biblioteca",
    "liquor_store": "Licorería",
    "local_government_office": "Oficina gubernamental",
    "locksmith": "Cerrajería",
    "lodging": "Alojamiento",
    "meal_delivery": "Delivery",
    "meal_takeaway": "Comida para llevar",
    "mosque": "Mezquita",
    "movie_theater": "Cine",
    "moving_company": "Mudanzas",
    "museum": "Museo",
    "night_club": "Boliche",
    "painter": "Pintor",
    "park": "Parque",
    "parking": "Estacionamiento",
    "pet_store": "Tienda de mascotas",
    "pharmacy": "Farmacia",
    "physiotherapist": "Fisioterapeuta",
    "plumber": "Plomero",
    "point_of_interest": "Punto de interés",
    "police": "Policía",
    "post_office": "Correo",
    "primary_school": "Escuela primaria",
    "real_estate_agency": "Inmobiliaria",
    "restaurant": "Restaurante",
    "roofing_contractor": "Techista",
    "rv_park": "Camping RV",
    "school": "Escuela",
    "secondary_school": "Escuela secundaria",
    "shoe_store": "Zapatería",
    "shopping_mall": "Shopping",
    "spa": "Spa",
    "stadium": "Estadio",
    "storage": "Depósito",
    "store": "Tienda",
    "subway_station": "Estación de subte",
    "supermarket": "Supermercado",
    "synagogue": "Sinagoga",
    "taxi_stand": "Parada de taxi",
    "tourist_attraction": "Atracción turística",
    "train_station": "Estación de tren",
    "transit_station": "Estación de transporte",
    "travel_agency": "Agencia de viajes",
    "university": "Universidad",
    "veterinary_care": "Veterinaria",
    "zoo": "Zoológico",
}

SKIP_TYPES = {
    "point_of_interest", "establishment", "food", "health",
    "store", "place_of_worship", "finance", "general_contractor",
    "political", "premise", "subpremise",
}

COLUMNS = [
    # Identification
    ("nombre", "Nombre"),
    ("categoria_principal", "Categoría"),
    ("keyword", "Palabra clave"),
    ("estado_negocio", "Estado"),
    # Contact
    ("telefono", "Teléfono"),
    ("telefono_internacional", "Teléfono intl."),
    ("email", "Email"),
    ("sitio_web", "Sitio web"),
    ("tiene_telefono", "Tiene teléfono"),
    ("tiene_web", "Tiene web"),
    ("tiene_email", "Tiene email"),
    ("score_contacto", "Score contacto"),
    # Reputation
    ("calificacion", "Calificación"),
    ("total_calificaciones", "Total reseñas"),
    ("score_reputacion", "Score reputación"),
    ("nivel_precio", "Nivel precio"),
    # Location
    ("direccion_completa", "Dirección completa"),
    ("barrio", "Barrio"),
    ("localidad", "Localidad"),
    ("provincia", "Provincia"),
    ("pais", "País"),
    ("calle", "Calle"),
    ("numero", "Número"),
    ("codigo_postal", "Código postal"),
    # Schedule
    ("horarios", "Horarios"),
    ("descripcion", "Descripción"),
    # Links & metadata
    ("enlace_maps", "Google Maps"),
    ("territorio", "Territorio"),
    ("fecha_relevamiento", "Fecha"),
    ("tipos", "Tipos (Google)"),
    ("latitud", "Latitud"),
    ("longitud", "Longitud"),
    ("google_place_id", "Place ID"),
    ("foto_url", "Foto URL"),
]

HEADER_KEYS = [c[0] for c in COLUMNS]
HEADER_LABELS = {c[0]: c[1] for c in COLUMNS}


def _translate_type(t: str) -> str | None:
    if t in SKIP_TYPES:
        return None
    return GOOGLE_TYPE_ES.get(t, t.replace("_", " ").title())


def _main_category(tipos: list[str] | None) -> str:
    if not tipos:
        return "--"
    for t in tipos:
        translated = _translate_type(t)
        if translated:
            return translated
    return "--"


def _score_contacto(row: dict) -> int:
    score = 0
    if row.get("telefono"):
        score += 1
    if row.get("sitio_web"):
        score += 1
    if row.get("email"):
        score += 1
    return score


def _score_reputacion(row: dict) -> str:
    rating = row.get("calificacion")
    count = row.get("total_calificaciones")
    if not rating or not count:
        return "Sin datos"
    try:
        r = float(rating)
        c = int(count)
    except (ValueError, TypeError):
        return "Sin datos"
    if c < 5 or r < 3.0:
        return "Baja"
    if r >= 4.2 and c >= 50:
        return "Alta"
    if r >= 3.5 and c >= 10:
        return "Media"
    return "Baja"


def _enrich(row: dict, territorio: str) -> dict:
    enriched = dict(row)
    enriched["territorio"] = territorio
    enriched["fecha_relevamiento"] = row.get("fecha_relevamiento", date.today().isoformat())
    enriched["categoria_principal"] = _main_category(row.get("tipos"))
    enriched["tiene_telefono"] = "Sí" if row.get("telefono") else "No"
    enriched["tiene_web"] = "Sí" if row.get("sitio_web") else "No"
    enriched["tiene_email"] = "Sí" if row.get("email") else "No"
    enriched["score_contacto"] = _score_contacto(row)
    enriched["score_reputacion"] = _score_reputacion(row)
    if isinstance(enriched.get("tipos"), list):
        enriched["tipos"] = " | ".join(enriched["tipos"])
    return enriched


def _cell_value(row: dict, key: str) -> str:
    val = row.get(key)
    if val is None:
        return "--"
    return str(val).replace("\n", " ")


def _escape_csv(val) -> str:
    if val is None:
        return '"--"'
    s = str(val).replace("\n", " ").replace('"', '""')
    return f'"{s}"'


def generate_csv(rows: list[dict], territorio: str) -> str:
    enriched = [_enrich(r, territorio) for r in rows]
    lines = [";".join(_escape_csv(HEADER_LABELS[k]) for k in HEADER_KEYS)]

    for row in enriched:
        line = [_escape_csv(_cell_value(row, k)) for k in HEADER_KEYS]
        lines.append(";".join(line))

    return "﻿" + "\r\n".join(lines)


def generate_xlsx(rows: list[dict], territorio: str) -> bytes:
    enriched = [_enrich(r, territorio) for r in rows]

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(bottom=Side(style="thin", color="CBD5E1"))
    cell_font = Font(size=10)
    link_font = Font(size=10, color="2563EB", underline="single")
    green_font = Font(size=10, color="15803D", bold=True)
    red_font = Font(size=10, color="DC2626")
    amber_font = Font(size=10, color="D97706")
    gray_font = Font(size=10, color="94A3B8")
    green_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    red_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

    for col_idx, key in enumerate(HEADER_KEYS, 1):
        cell = ws.cell(row=1, column=col_idx, value=HEADER_LABELS[key])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    numeric_keys = {"calificacion", "total_calificaciones", "nivel_precio", "latitud", "longitud", "score_contacto"}
    link_keys = {"enlace_maps", "sitio_web"}

    for row_idx, row in enumerate(enriched, 2):
        for col_idx, key in enumerate(HEADER_KEYS, 1):
            val = _cell_value(row, key)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.font = cell_font

            if key in numeric_keys:
                try:
                    cell.value = float(val) if val != "--" else None
                except ValueError:
                    cell.value = val
            elif key in link_keys and val != "--":
                cell.value = val
                cell.font = link_font
                cell.hyperlink = val
            elif key in ("tiene_telefono", "tiene_web", "tiene_email"):
                cell.value = val
                if val == "Sí":
                    cell.font = green_font
                    cell.fill = green_fill
                else:
                    cell.font = red_font
                    cell.fill = red_fill
            elif key == "email" and val != "--":
                cell.value = val
                cell.font = link_font
                cell.hyperlink = f"mailto:{val}"
            elif key == "score_reputacion":
                cell.value = val
                if val == "Alta":
                    cell.font = green_font
                    cell.fill = green_fill
                elif val == "Media":
                    cell.font = amber_font
                elif val == "Baja":
                    cell.font = red_font
                else:
                    cell.font = gray_font
            elif key == "score_contacto":
                try:
                    cell.value = int(val) if val != "--" else 0
                except ValueError:
                    cell.value = val
                if cell.value >= 2:
                    cell.font = green_font
                elif cell.value == 1:
                    cell.font = amber_font
                else:
                    cell.font = red_font
            else:
                cell.value = val if val != "--" else None

    for col_idx in range(1, len(HEADER_KEYS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row_idx in range(2, min(len(enriched) + 2, 52)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[col_letter].width = max_len + 3

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADER_KEYS))}{len(enriched) + 1}"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(territorio: str, ext: str) -> str:
    fecha = date.today().isoformat()
    terr = "".join(c if c.isalnum() or c in "_-" else "_" for c in territorio)[:30]
    return f"ProspectoAI_{terr}_{fecha}.{ext}"
